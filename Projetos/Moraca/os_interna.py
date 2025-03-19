import os
import re
import sqlite3
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

def criar_os_interna(numero_os, pasta_destino=None):
    """
    Cria um arquivo Excel formatado para a OS Interna.
    
    Args:
        numero_os (str): Número da Ordem de Serviço
        pasta_destino (str, optional): Caminho onde o arquivo será salvo. Se None, usa o caminho padrão.
    
    Returns:
        str: Caminho completo do arquivo salvo
    """
    print(f"Iniciando criação de OS Interna para número: '{numero_os}'")
    
    # Formatação do número da OS (se ainda não estiver no formato padrão)
    if not numero_os.startswith("OS "):
        numeros = re.findall(r'\d+', numero_os)
        
        if numeros:
            numero_completo = ''.join(numeros)
            if len(numero_completo) >= 5:
                ano = numero_completo[:2]
                sequencia = numero_completo[-3:].zfill(3)  # Garantir 3 dígitos
            else:
                ano = datetime.now().strftime("%y")
                sequencia = numero_completo.zfill(3)  # Garantir 3 dígitos
                
            numero_os_formatado = f"OS {ano} {sequencia}"
        else:
            ano = datetime.now().strftime("%y")
            numero_os_formatado = f"OS {ano} 001"
    else:
        numero_os_formatado = numero_os
    
    # Formatos alternativos para consulta no banco
    numero_os_sem_espacos = numero_os_formatado.replace(" ", "")  # OSxxyyy
    numero_os_original = numero_os
            
    print(f"Número formatado: '{numero_os_formatado}'")
    print(f"Número sem espaços: '{numero_os_sem_espacos}'")
            
    # Definir pasta de destino padrão se não for fornecida
    if pasta_destino is None:
        pasta_destino = os.path.join("MORACA", "MORACA1", "DOCUMENTOS", "tecnico", "interna")
    
    # Criar diretórios se não existirem
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
        
    # Criar nome de arquivo baseado no número da OS formatado
    numero_os_limpo = numero_os_formatado.replace(" ", "_").replace("/", "-").replace("-", "_")
    nome_arquivo = f"OS_Interna_{numero_os_limpo}.xlsx"
    caminho_arquivo = os.path.join(pasta_destino, nome_arquivo)
    
    # Verificar se o arquivo já existe
    if os.path.exists(caminho_arquivo):
        print(f"Arquivo já existe em: {caminho_arquivo}")
        return caminho_arquivo
        
    # Buscar informações da OS do banco de dados
    conn = sqlite3.connect('moraca.db')
    conn.row_factory = sqlite3.Row  # Para acessar as colunas pelo nome
    c = conn.cursor()
    
    # Tentar diferentes formatos para encontrar a OS no banco
    print(f"Buscando OS no banco de dados...")
    formatos_para_testar = [
        numero_os_formatado,  # "OS xx yyy"
        numero_os_sem_espacos,  # "OSxxttt"
        numero_os_original,  # formato original passado para a função
    ]
    
    os_data = None
    
    for formato in formatos_para_testar:
        print(f"Tentando consulta com formato: '{formato}'")
        c.execute("SELECT * FROM os WHERE numero = ?", (formato,))
        os_data = c.fetchone()
        if os_data:
            print(f"OS encontrada com formato: '{formato}'")
            break
    
    if not os_data:
        print(f"OS não encontrada. Listando algumas OS disponíveis:")
        c.execute("SELECT numero FROM os LIMIT 5")
        os_disponiveis = c.fetchall()
        for os_disp in os_disponiveis:
            print(f"  - {os_disp[0]}")
        
        conn.close()
        return None
    
    # Criar uma nova planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"OS {numero_os_formatado}"
    
    # Definir estilos
    titulo_font = Font(name='Arial', size=16, bold=True)
    subtitulo_font = Font(name='Arial', size=14, bold=True)
    cabecalho_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Bordas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Estilo de linha pontilhada horizontal para tabela
    dotted_border = Border(
        bottom=Side(style='dotted')
    )
    
    # Cores para células
    cinza_claro = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    laranja_claro = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')  # Cor salmão/laranja claro
    
    # Configurar larguras das colunas seguindo a imagem de referência
    ws.column_dimensions['A'].width = 21
    ws.column_dimensions['B'].width = 30  # Aumentando a largura para caber a frase completa
    ws.column_dimensions['C'].width = 21
    ws.column_dimensions['D'].width = 21
    ws.column_dimensions['E'].width = 8
    
    # Ajustar altura das linhas iniciais
    ws.row_dimensions[1].height = 25  # Altura linha do título O.S.I.
    ws.row_dimensions[2].height = 25  # Altura linha do número e prazo
    
    # ===== Título e cabeçalho =====
    # O.S.I. - Título principal
    ws['A1'] = "O.S.I."
    ws['A1'].font = titulo_font
    ws['A1'].alignment = left_align
    
    # Número da OS formatado com o padrão correto (ex: OS25102)
    # Extrair o número formatado da OS
    if numero_os_sem_espacos.startswith("OS"):
        ws['A2'] = numero_os_sem_espacos  # Usar o formato OSxxxxx
    else:
        # Se não tiver o formato esperado, usar o número original
        ws['A2'] = numero_os
    ws['A2'].font = titulo_font
    ws['A2'].alignment = left_align
    
    # PRAZO DE ENTREGA FINAL
    ws['B2'] = "PRAZO DE ENTREGA FINAL:"
    ws['B2'].font = cabecalho_font
    ws['B2'].alignment = right_align
    
    # Campo para data __/__/____
    ws['D2'] = "__/__/____"
    ws['D2'].font = cabecalho_font
    ws['D2'].alignment = center_align
    
    # ===== Campos de informações =====
    row = 4
    
    # Cliente
    ws['A4'] = "Cliente:"
    ws['A4'].font = normal_font
    ws['A4'].alignment = right_align
    
    # Valor cliente
    cliente = os_data['cliente'] if os_data['cliente'] is not None else ''
    ws.merge_cells('B4:E4')
    ws['B4'] = cliente
    ws['B4'].font = normal_font
    ws['B4'].alignment = left_align
    
    # Equipamento
    row += 1
    ws['A5'] = "Equipamento :"
    ws['A5'].font = normal_font
    ws['A5'].alignment = right_align
    
    # Valor equipamento
    equipamento = os_data['maquina'] if os_data['maquina'] is not None else ''
    ws.merge_cells('B5:E5')
    ws['B5'] = equipamento
    ws['B5'].font = normal_font
    ws['B5'].alignment = left_align
    
    # Número de Série
    row += 1
    ws['A6'] = "Número de Série:"
    ws['A6'].font = normal_font
    ws['A6'].alignment = right_align
    
    # Valor número de série
    numero_serie = os_data['numero_serie'] if os_data['numero_serie'] is not None else ''
    ws.merge_cells('B6:E6')
    ws['B6'] = numero_serie
    ws['B6'].font = normal_font
    ws['B6'].alignment = left_align
    
    # Data de abertura O.S.I.
    row += 1
    ws['A7'] = "Data de abertura O.S.:"
    ws['A7'].font = normal_font
    ws['A7'].alignment = right_align
    
    # Valor data
    data = os_data['data'] if os_data['data'] is not None else datetime.now().strftime("%d/%m/%Y")
    ws.merge_cells('B7:E7')
    ws['B7'] = data
    ws['B7'].font = normal_font
    ws['B7'].alignment = left_align
    
    # Horário de abertura O.S.I.
    row += 1
    ws['A8'] = "Horário de abertura O.S.:"
    ws['A8'].font = normal_font
    ws['A8'].alignment = right_align
    
    # Valor horário
    ws.merge_cells('B8:E8')
    ws['B8'] = datetime.now().strftime("%H:%M")  # Hora atual
    ws['B8'].font = normal_font
    ws['B8'].alignment = left_align
    
    # ===== Tabela de tarefas =====
    row = 10
    
    # Cabeçalho tabela
    ws.merge_cells('A10:C10')
    ws['A10'] = "TAREFA A EXECUTAR"
    ws['A10'].font = cabecalho_font
    ws['A10'].alignment = center_align
    ws['A10'].fill = cinza_claro
    ws['A10'].border = thin_border
    
    ws['D10'] = "PRAZO PREVISTO"
    ws['D10'].font = cabecalho_font
    ws['D10'].alignment = center_align
    ws['D10'].fill = cinza_claro
    ws['D10'].border = thin_border
    
    ws['E10'] = "OK"
    ws['E10'].font = cabecalho_font
    ws['E10'].alignment = center_align
    ws['E10'].fill = cinza_claro
    ws['E10'].border = thin_border
    
    # Aplicando bordas no bloco TAREFA A EXECUTAR e PRAZO PREVISTO
    # Primeiro definir as bordas para o bloco A11:C23 (TAREFA A EXECUTAR)
    for i in range(11, 24):
        # Mesclando as células
        ws.merge_cells(f'A{i}:C{i}')
        
        if i == 11:  # Primeira linha - borda superior e laterais
            ws[f'A{i}'].border = Border(left=Side(style='thin'), top=Side(style='thin'), right=Side(style='thin'))
        elif i == 23:  # Última linha - borda inferior e laterais
            ws[f'A{i}'].border = Border(left=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
        else:  # Linhas do meio - apenas bordas laterais
            ws[f'A{i}'].border = Border(left=Side(style='thin'), right=Side(style='thin'))
        
        # Adicionando as bordas na coluna OK (E)
        ws[f'E{i}'].border = thin_border
    
    # Adicionando a borda ao redor do bloco D11:D23 (PRAZO PREVISTO)
    for i in range(11, 24):
        if i == 11:  # Primeira linha - borda superior e laterais
            ws[f'D{i}'].border = Border(left=Side(style='thin'), top=Side(style='thin'), right=Side(style='thin'))
        elif i == 23:  # Última linha - borda inferior e laterais
            ws[f'D{i}'].border = Border(left=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
        else:  # Linhas do meio - apenas bordas laterais
            ws[f'D{i}'].border = Border(left=Side(style='thin'), right=Side(style='thin'))
    
    # ===== Seção de observações =====
    row = 24
    ws.merge_cells('A24:E24')
    ws['A24'] = "OBSERVAÇÕES:"
    ws['A24'].font = cabecalho_font
    ws['A24'].alignment = center_align
    ws['A24'].fill = cinza_claro
    ws['A24'].border = thin_border
    
    # Aplicando bordas na área de observações
    # Mesclando todas as células e aplicando bordas laterais
    for i in range(25, 36):
        ws.merge_cells(f'A{i}:E{i}')
        
        if i == 25:  # Primeira linha após o cabeçalho
            ws[f'A{i}'] = "Quando for trocar tubo favor:"
            ws[f'A{i}'].border = Border(left=Side(style='thin'), right=Side(style='thin'))
        elif i >= 26 and i <= 28:  # Linhas com marcadores
            texto_idx = i - 26
            texto_observacoes = [
                "* Descrição do tubo:",
                "* N/S° do tubo:",
                "* Tamanho do foco::"
            ]
            if texto_idx < len(texto_observacoes):
                ws[f'A{i}'] = texto_observacoes[texto_idx]
            ws[f'A{i}'].border = Border(left=Side(style='thin'), right=Side(style='thin'))
        elif i == 35:  # Última linha - com borda inferior
            ws[f'A{i}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
        else:  # Demais linhas - apenas bordas laterais
            ws[f'A{i}'].border = Border(left=Side(style='thin'), right=Side(style='thin'))
        
        ws[f'A{i}'].font = normal_font
        ws[f'A{i}'].alignment = left_align
    
    # Definir configurações de impressão
    ws.print_area = 'A1:E36'
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    
    # Adicionar borda ao redor de todo o documento
    # Criar uma borda simples que vai envolver todo o documento
    outer_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Aplicar bordas às bordas externas do documento
    # Primeiro, definimos as células dos cantos
    ws['A1'].border = Border(
        left=Side(style='thin'), 
        top=Side(style='thin'), 
        right=None, 
        bottom=None
    )
    
    ws['E1'].border = Border(
        left=None, 
        top=Side(style='thin'), 
        right=Side(style='thin'), 
        bottom=None
    )
    
    ws['A36'].border = Border(
        left=Side(style='thin'), 
        top=None, 
        right=None, 
        bottom=Side(style='thin')
    )
    
    ws['E36'].border = Border(
        left=None, 
        top=None, 
        right=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Depois, definimos as bordas das extremidades
    # Borda superior (B1-D1)
    for col in ['B', 'C', 'D']:
        ws[f'{col}1'].border = Border(top=Side(style='thin'))
    
    # Borda inferior (B36-D36)
    for col in ['B', 'C', 'D']:
        ws[f'{col}36'].border = Border(bottom=Side(style='thin'))
    
    # Borda esquerda (A2-A35)
    for row in range(2, 36):
        ws[f'A{row}'].border = Border(left=Side(style='thin'))
    
    # Borda direita (E2-E35)
    for row in range(2, 36):
        ws[f'E{row}'].border = Border(right=Side(style='thin'))
    
    # Salvar o arquivo
    try:
        wb.save(caminho_arquivo)
        print(f"OS Interna salva com sucesso em: {caminho_arquivo}")
        conn.close()
        return caminho_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo: {e}")
        conn.close()
        return None

# Para testes
if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        numero_os = sys.argv[1]
        print(f"Criando OS Interna para número: {numero_os}")
        caminho = criar_os_interna(numero_os)
        print(f"Arquivo salvo em: {caminho}")
    else:
        print("Por favor, forneça o número da OS como argumento.") 