import os
import re
import sqlite3
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

def criar_os_visita_tecnica(numero_os, pasta_destino=None):
    """
    Cria um arquivo Excel formatado para a OS de visita técnica.
    
    Args:
        numero_os (str): Número da Ordem de Serviço
        pasta_destino (str, optional): Caminho onde o arquivo será salvo. Se None, usa o caminho padrão.
    
    Returns:
        str: Caminho completo do arquivo salvo
    """
    print(f"Iniciando criação de OS Visita Técnica para número: '{numero_os}'")
    
    # Formatação do número da OS (se ainda não estiver no formato padrão)
    if not numero_os.startswith("OS "):
        numeros = re.findall(r'\d+', numero_os)
        
        if numeros:
            numero_completo = ''.join(numeros)
            if len(numero_completo) >= 5:
                ano = numero_completo[:2]
                sequencia = numero_completo[-3:].zfill(3)  # Garantir 3 dígitos
            else:
                ano = datetime.now().strftime("%y")
                sequencia = numero_completo.zfill(3)  # Garantir 3 dígitos
                
            numero_os_formatado = f"OS {ano} {sequencia}"
        else:
            ano = datetime.now().strftime("%y")
            numero_os_formatado = f"OS {ano} 001"
    else:
        numero_os_formatado = numero_os
    
    # Formatos alternativos para consulta no banco
    numero_os_sem_espacos = numero_os_formatado.replace(" ", "")  # OSxxyyy
    numero_os_original = numero_os
            
    print(f"Número formatado: '{numero_os_formatado}'")
    print(f"Número sem espaços: '{numero_os_sem_espacos}'")
            
    # Definir pasta de destino padrão se não for fornecida
    if pasta_destino is None:
        pasta_destino = os.path.join("MORACA", "MORACA1", "DOCUMENTOS", "tecnico", "visitas")
    
    # Criar diretórios se não existirem
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
        
    # Criar nome de arquivo baseado no número da OS formatado
    numero_os_limpo = numero_os_formatado.replace(" ", "_").replace("/", "-").replace("-", "_")
    nome_arquivo = f"OS_Visita_Tecnica_{numero_os_limpo}.xlsx"
    caminho_arquivo = os.path.join(pasta_destino, nome_arquivo)
    
    # Verificar se o arquivo já existe
    if os.path.exists(caminho_arquivo):
        print(f"Arquivo já existe em: {caminho_arquivo}")
        return caminho_arquivo
        
    # Buscar informações da OS do banco de dados
    conn = sqlite3.connect('moraca.db')
    conn.row_factory = sqlite3.Row  # Para acessar as colunas pelo nome
    c = conn.cursor()
    
    # Tentar diferentes formatos para encontrar a OS no banco
    print(f"Buscando OS no banco de dados...")
    formatos_para_testar = [
        numero_os_formatado,  # "OS xx yyy"
        numero_os_sem_espacos,  # "OSxxttt"
        numero_os_original,  # formato original passado para a função
    ]
    
    os_data = None
    
    for formato in formatos_para_testar:
        print(f"Tentando consulta com formato: '{formato}'")
        c.execute("SELECT * FROM os WHERE numero = ?", (formato,))
        os_data = c.fetchone()
        if os_data:
            print(f"OS encontrada com formato: '{formato}'")
            break
    
    if not os_data:
        print(f"OS não encontrada. Listando algumas OS disponíveis:")
        c.execute("SELECT numero FROM os LIMIT 5")
        os_disponiveis = c.fetchall()
        for os_disp in os_disponiveis:
            print(f"  - {os_disp[0]}")
        
        conn.close()
        return None
    
    # Criar uma nova planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"OS {numero_os_formatado}"
    
    # Definir estilos
    titulo_font = Font(name='Arial', size=24, bold=True)
    subtitulo_font = Font(name='Arial', size=14, bold=True)
    cabecalho_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Bordas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    medium_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    # Cores para células
    cinza_claro = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    # Configurar larguras das colunas
    ws.column_dimensions['A'].width = 30  # Aumentar largura da coluna A
    ws.column_dimensions['B'].width = 25  # Aumentar um pouco a coluna B para valores
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    
    # Ajustar altura das linhas
    ws.row_dimensions[1].height = 30  # Cabeçalho
    ws.row_dimensions[2].height = 25  # Linhas do número da OS
    ws.row_dimensions[3].height = 25
    ws.row_dimensions[4].height = 25
    
    # Ajustar altura das linhas de informações
    for r in range(5, 16):
        ws.row_dimensions[r].height = 18  # Altura uniforme para os dados
    
    # Definir área de impressão
    ws.print_area = 'A1:E35'  # Ajustar conforme necessário
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    
    # Título principal
    ws.merge_cells('A1:E1')
    ws['A1'] = "Ordem de Serviço"
    ws['A1'].font = titulo_font
    ws['A1'].alignment = center_align
    ws['A1'].border = thin_border
    
    # Número da OS (em destaque) - Ajustando espaçamento
    ws.merge_cells('A2:B4')
    numero_formatado = f"{numero_os_formatado.replace('OS ', '')}-1"  # Formato: XX.XXXX-1
    ws['A2'] = f"OS\n{numero_formatado}"
    ws['A2'].font = Font(name='Arial', size=36, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # Removendo borda do número da OS
    ws['A2'].border = None
    
    # Células para o logo (dividindo em duas áreas)
    ws.merge_cells('C2:C4')
    # Removendo borda
    ws['C2'].border = None
    
    ws.merge_cells('D2:E4')
    # Removendo borda
    ws['D2'].border = None
    
    # Adicionar logo Moraca
    logo_path = os.path.join("assets", "logo.png")
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            # Ajustar tamanho
            max_width = 200
            max_height = 80
            width_ratio = max_width / img.width
            height_ratio = max_height / img.height
            ratio = min(width_ratio, height_ratio)
            img.width = int(img.width * ratio)
            img.height = int(img.height * ratio)
            
            # Adicionar imagem na célula D2 (células mescladas D2:E4)
            ws.add_image(img, 'D2')
        except Exception as e:
            print(f"Erro ao adicionar logo: {e}")
    
    # Dados do equipamento - Reorganizando o layout das linhas 5 a 15
    row = 5
    # Não mesclar células para separar rótulos e valores
    ws['A5'] = "Equipamento :"
    ws['A5'].font = normal_font
    ws['A5'].alignment = left_align
    # Removendo borda
    ws['A5'].border = None
    
    # Valor em B5
    ws.merge_cells(f'B5:E5')
    ws['B5'] = f"{os_data['maquina']}"
    ws['B5'].font = normal_font
    ws['B5'].alignment = left_align
    # Removendo borda
    ws['B5'].border = None
    ws['C5'].border = None
    ws['D5'].border = None
    ws['E5'].border = None
    
    row += 1
    ws['A6'] = "Nº Patrimônio:"
    ws['A6'].font = normal_font
    ws['A6'].alignment = left_align
    # Removendo borda
    ws['A6'].border = None
    
    # Valor em B6
    ws.merge_cells(f'B6:E6')
    patrimonio = os_data['patrimonio'] if os_data['patrimonio'] is not None else ''
    ws['B6'] = f"{patrimonio}"
    ws['B6'].font = normal_font
    ws['B6'].alignment = left_align
    # Removendo borda
    ws['B6'].border = None
    ws['C6'].border = None
    ws['D6'].border = None
    ws['E6'].border = None
    
    row += 1
    ws['A7'] = "Nº Série:"
    ws['A7'].font = normal_font
    ws['A7'].alignment = left_align
    # Removendo borda
    ws['A7'].border = None
    
    # Valor em B7
    ws.merge_cells(f'B7:E7')
    numero_serie = os_data['numero_serie'] if os_data['numero_serie'] is not None else ''
    ws['B7'] = f"{numero_serie}"
    ws['B7'].font = normal_font
    ws['B7'].alignment = left_align
    # Removendo borda
    ws['B7'].border = None
    ws['C7'].border = None
    ws['D7'].border = None
    ws['E7'].border = None
    
    # Local
    row += 1
    ws['A8'] = "Local de Instalação :"
    ws['A8'].font = normal_font
    ws['A8'].alignment = left_align
    # Removendo borda
    ws['A8'].border = None
    
    # Valor em B8
    ws.merge_cells(f'B8:E8')
    local = os_data['local'] if os_data['local'] is not None else ''
    ws['B8'] = f"{local}"
    ws['B8'].font = normal_font
    ws['B8'].alignment = left_align
    # Removendo borda
    ws['B8'].border = None
    ws['C8'].border = None
    ws['D8'].border = None
    ws['E8'].border = None
    
    # Endereço e CEP
    row += 1
    ws['A9'] = "Endereço:"
    ws['A9'].font = normal_font
    ws['A9'].alignment = left_align
    # Removendo borda
    ws['A9'].border = None
    
    # Valor em B9
    ws.merge_cells(f'B9:E9')
    endereco = os_data['endereco'] if os_data['endereco'] is not None else ''
    ws['B9'] = f"{endereco}"
    ws['B9'].font = normal_font
    ws['B9'].alignment = left_align
    # Removendo borda
    ws['B9'].border = None
    ws['C9'].border = None
    ws['D9'].border = None
    ws['E9'].border = None
    
    row += 1
    ws['A10'] = "Cidade:"
    ws['A10'].font = normal_font
    ws['A10'].alignment = left_align
    # Removendo borda
    ws['A10'].border = None
    
    # Valor em B10
    ws.merge_cells(f'B10:E10')
    # Para o campo cidade, usar valor do banco de dados sem valor padrão
    cidade = ""  # Valor vazio por padrão
    
    # Verificar se a coluna existe nos metadados da consulta
    colunas = [column[0] for column in c.description]
    if 'cidade' in colunas:
        cidade = os_data['cidade'] if os_data['cidade'] is not None else ""
    
    # Garantir que não há valor padrão
    ws['B10'] = cidade
    ws['B10'].font = normal_font
    ws['B10'].alignment = left_align
    # Removendo borda
    ws['B10'].border = None
    ws['C10'].border = None
    ws['D10'].border = None
    ws['E10'].border = None
    
    row += 1
    ws['A11'] = "Cep:"
    ws['A11'].font = normal_font
    ws['A11'].alignment = left_align
    # Removendo borda
    ws['A11'].border = None
    
    # Valor em B11
    ws.merge_cells(f'B11:E11')
    cep = os_data['cep'] if os_data['cep'] is not None else ''
    ws['B11'] = f"{cep}"
    ws['B11'].font = normal_font
    ws['B11'].alignment = left_align
    # Removendo borda
    ws['B11'].border = None
    ws['C11'].border = None
    ws['D11'].border = None
    ws['E11'].border = None
    
    # Telefone
    row += 1
    ws['A12'] = "Telefone:"
    ws['A12'].font = normal_font
    ws['A12'].alignment = left_align
    # Removendo borda
    ws['A12'].border = None
    
    # Valor em B12
    ws.merge_cells(f'B12:E12')
    telefone = os_data['telefone'] if os_data['telefone'] is not None else ''
    ws['B12'] = f"{telefone}"
    ws['B12'].font = normal_font
    ws['B12'].alignment = left_align
    # Removendo borda
    ws['B12'].border = None
    ws['C12'].border = None
    ws['D12'].border = None
    ws['E12'].border = None
    
    # Solicitante
    row += 1
    ws['A13'] = "Solicitante :"
    ws['A13'].font = normal_font
    ws['A13'].alignment = left_align
    # Removendo borda
    ws['A13'].border = None
    
    # Valor em B13
    ws.merge_cells(f'B13:E13')
    contato_nome = os_data['contato_nome'] if os_data['contato_nome'] is not None else ''
    contato_tel1 = os_data['contato_telefone1'] if os_data['contato_telefone1'] is not None else ''
    ws['B13'] = f"{contato_nome} {contato_tel1}"
    ws['B13'].font = normal_font
    ws['B13'].alignment = left_align
    # Removendo borda
    ws['B13'].border = None
    ws['C13'].border = None
    ws['D13'].border = None
    ws['E13'].border = None
    
    # Contato
    row += 1
    ws['A14'] = "Contato :"
    ws['A14'].font = normal_font
    ws['A14'].alignment = right_align
    # Removendo borda
    ws['A14'].border = None
    
    # Valor em B14
    ws.merge_cells(f'B14:E14')
    ws['B14'] = f"{contato_nome} {contato_tel1}"
    ws['B14'].font = normal_font
    ws['B14'].alignment = right_align
    # Removendo borda
    ws['B14'].border = None
    ws['C14'].border = None
    ws['D14'].border = None
    ws['E14'].border = None
    
    # Data
    row += 1
    ws['A15'] = "Data :"
    ws['A15'].font = normal_font
    ws['A15'].alignment = right_align
    # Removendo borda
    ws['A15'].border = None
    
    # Valor em B15
    ws.merge_cells(f'B15:E15')
    data_str = os_data['data'] if os_data['data'] is not None else datetime.now().strftime("%d/%m/%Y")
    ws['B15'] = f"{data_str}"
    ws['B15'].font = normal_font
    ws['B15'].alignment = left_align
    # Removendo borda
    ws['B15'].border = None
    ws['C15'].border = None
    ws['D15'].border = None
    ws['E15'].border = None
    
    # Seções com fundo cinza
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "TESTE INICIAL DE AVALIAÇÃO:"
    ws[f'A{row}'].font = cabecalho_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = cinza_claro
    ws[f'A{row}'].border = thin_border
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = ""  # Área para preenchimento
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = left_align
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 50  # Altura maior para campo de preenchimento
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "PROBLEMAS APRESENTADOS:"
    ws[f'A{row}'].font = cabecalho_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = cinza_claro
    ws[f'A{row}'].border = thin_border
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = ""  # Área para preenchimento
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = left_align
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 50  # Altura maior para campo de preenchimento
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "SOLICITAÇÃO DO CLIENTE:"
    ws[f'A{row}'].font = cabecalho_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = cinza_claro
    ws[f'A{row}'].border = thin_border
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = ""  # Área para preenchimento
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = left_align
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 50  # Altura maior para campo de preenchimento
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "SERVIÇOS REALIZADOS:"
    ws[f'A{row}'].font = cabecalho_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = cinza_claro
    ws[f'A{row}'].border = thin_border
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = ""  # Área para preenchimento
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = left_align
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 140  # Área grande para descrição dos serviços
    
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "Obs :"
    ws[f'A{row}'].font = cabecalho_font
    ws[f'A{row}'].alignment = left_align
    ws[f'A{row}'].border = thin_border
    
    # Adicionando espaço para observações
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = ""  # Área para observações
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = left_align
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 40  # Altura para observações
    
    # Assinaturas
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Responsável :"
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = left_align
    ws[f'A{row}'].border = thin_border
    
    ws.merge_cells(f'C{row}:E{row}')
    ws[f'C{row}'] = ""  # Espaço para preenchimento do responsável
    ws[f'C{row}'].border = thin_border
    
    # Linha para horários
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Horário Entrada :"
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = left_align
    ws[f'A{row}'].border = thin_border
    
    ws[f'C{row}'] = ""  # Espaço para hora de entrada
    ws[f'C{row}'].border = thin_border
    
    ws.merge_cells(f'D{row}:E{row}')
    ws[f'D{row}'] = "Horário Saída :"
    ws[f'D{row}'].font = normal_font
    ws[f'D{row}'].alignment = left_align
    ws[f'D{row}'].border = thin_border
    
    # Linha para data
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Data :"
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = left_align
    ws[f'A{row}'].border = thin_border
    
    ws.merge_cells(f'C{row}:E{row}')
    ws[f'C{row}'] = "    /    /    "
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = center_align
    ws[f'C{row}'].border = thin_border
    
    # Espaço antes das assinaturas
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = ""
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 15
    
    # Assinaturas - empresa e cliente
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Moraca Ltda"
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = thin_border
    
    ws[f'C{row}'].border = thin_border
    
    ws.merge_cells(f'D{row}:E{row}')
    ws[f'D{row}'] = "Cliente"
    ws[f'D{row}'].font = normal_font
    ws[f'D{row}'].alignment = center_align
    ws[f'D{row}'].border = thin_border
    
    # Linhas para assinatura
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "____________________________"
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = thin_border
    
    ws[f'C{row}'].border = thin_border
    
    ws.merge_cells(f'D{row}:E{row}')
    ws[f'D{row}'] = "____________________________"
    ws[f'D{row}'].font = normal_font
    ws[f'D{row}'].alignment = center_align
    ws[f'D{row}'].border = thin_border
    
    # Rodapé
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "E-mail moraca@moraca.com.br - Razão Social: Moraca ltda."
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = thin_border
    
    # Aplicar bordas às células em todo o documento
    def aplicar_bordas(worksheet, ultima_linha):
        """Aplica bordas a todas as células usadas na planilha"""
        # Primeiro, limpar todas as bordas (exceto a linha 1 que tem o título)
        for row_idx in range(2, ultima_linha + 1):
            for col_idx in range(1, 6):  # Colunas A a E
                col_letter = get_column_letter(col_idx)
                cell = worksheet[f"{col_letter}{row_idx}"]
                if hasattr(cell, "border"):
                    # Remover qualquer borda definida
                    cell.border = None
        
        # Aplicar bordas apenas a partir da linha 16 em diante (depois das informações de cabeçalho)
        for row_idx in range(16, ultima_linha + 1):
            for col_idx in range(1, 6):  # Colunas A a E
                col_letter = get_column_letter(col_idx)
                cell = worksheet[f"{col_letter}{row_idx}"]
                cell.border = thin_border
        
        # Manter a borda do título
        worksheet['A1'].border = thin_border
        worksheet['B1'].border = thin_border
        worksheet['C1'].border = thin_border
        worksheet['D1'].border = thin_border
        worksheet['E1'].border = thin_border
    
    # Aplicar bordas até a última linha usada
    aplicar_bordas(ws, row)
    
    # Salvar o arquivo
    try:
        wb.save(caminho_arquivo)
        print(f"OS de Visita Técnica salva com sucesso em: {caminho_arquivo}")
        conn.close()
        return caminho_arquivo
    except Exception as e:
        print(f"Erro ao salvar arquivo: {e}")
        conn.close()
        return None

# Para testes
if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        numero_os = sys.argv[1]
        print(f"Criando OS de Visita Técnica para número: {numero_os}")
        caminho = criar_os_visita_tecnica(numero_os)
        print(f"Arquivo salvo em: {caminho}")
    else:
        print("Por favor, forneça o número da OS como argumento.") 