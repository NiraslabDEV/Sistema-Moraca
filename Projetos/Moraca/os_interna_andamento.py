import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import re
import locale
from openpyxl.drawing.image import Image as XLImage

# Configurar o locale para português do Brasil
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')
    except:
        pass  # Se falhar, usará o locale padrão

# Função para formatar a data em português independentemente do locale
def formatar_data_pt_br(data):
    """Formata a data em português, independentemente do locale."""
    dias_semana = {
        0: "Segunda-feira",
        1: "Terça-feira",
        2: "Quarta-feira",
        3: "Quinta-feira",
        4: "Sexta-feira",
        5: "Sábado",
        6: "Domingo"
    }
    meses = {
        1: "janeiro",
        2: "fevereiro",
        3: "março",
        4: "abril",
        5: "maio",
        6: "junho",
        7: "julho",
        8: "agosto",
        9: "setembro",
        10: "outubro",
        11: "novembro",
        12: "dezembro"
    }
    
    dia_semana = dias_semana[data.weekday()]
    dia = data.day
    mes = meses[data.month]
    ano = data.year
    
    return f"{dia_semana}, {dia} de {mes} de {ano}"

def criar_os_interna_andamento(numero_os, pasta_destino=None):
    """
    Cria um arquivo Excel com duas planilhas para a OS interna de andamento.
    
    Args:
        numero_os (str): Número da Ordem de Serviço
        pasta_destino (str, optional): Caminho onde o arquivo será salvo. Se None, usa o caminho padrão.
    
    Returns:
        str: Caminho completo do arquivo salvo
    """
    # Formatação do número da OS (se ainda não estiver no formato padrão)
    if not numero_os.startswith("OS "):
        # Extrair os números do string, caso seja algo como "OSI252" ou outro formato
        numeros = re.findall(r'\d+', numero_os)
        
        if numeros:
            numero_completo = ''.join(numeros)
            # Se tiver 5 dígitos ou mais, assumimos que os 2 primeiros são do ano
            if len(numero_completo) >= 5:
                ano = numero_completo[:2]
                sequencia = numero_completo[-3:].zfill(3)  # Garantir 3 dígitos, preenchendo com zeros
            else:
                # Caso não tenha dígitos suficientes, usar o ano atual
                ano = datetime.now().strftime("%y")
                sequencia = numero_completo.zfill(3)  # Garantir 3 dígitos
                
            # Formatar no padrão correto
            numero_os_formatado = f"OS {ano} {sequencia}"
        else:
            # Se não tiver números, criar um padrão com o ano atual e sequência 001
            ano = datetime.now().strftime("%y")
            numero_os_formatado = f"OS {ano} 001"
    else:
        numero_os_formatado = numero_os
            
    # Definir pasta de destino padrão se não for fornecida
    if pasta_destino is None:
        pasta_destino = os.path.join("MORACA", "MORACA1", "DOCUMENTOS", "tecnico", "andamento")
    
    # Criar diretórios se não existirem
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
        
    # Criar nome de arquivo baseado no número da OS formatado
    numero_os_limpo = numero_os_formatado.replace(" ", "_").replace("/", "-").replace("-", "_")
    nome_arquivo = f"OS_{numero_os_limpo}.xlsx"
    caminho_arquivo = os.path.join(pasta_destino, nome_arquivo)
    
    # Verificar se o arquivo já existe
    if os.path.exists(caminho_arquivo):
        print(f"Arquivo já existe em: {caminho_arquivo}")
        return caminho_arquivo
        
    # Se não existe, criar um novo arquivo
    print(f"Criando novo arquivo em: {caminho_arquivo}")
    
    # Criar uma nova planilha
    wb = openpyxl.Workbook()
    
    # Definir estilos comuns
    titulo_font = Font(name='Arial', size=16, bold=True)
    cabecalho_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Bordas para as células
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Borda pontilhada para células de preenchimento (simulada com borda fina)
    dotted_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='dotted')
    )
    
    # Fundo cinza claro para cabeçalhos
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    # ============== PÁGINA 1: ANDAMENTO ==============
    # Renomear a primeira planilha
    ws1 = wb.active
    ws1.title = "Andamento"
    
    # Configurar larguras das colunas
    ws1.column_dimensions['A'].width = 25  # Funcionário
    ws1.column_dimensions['B'].width = 12  # Data
    ws1.column_dimensions['C'].width = 10  # H. Início
    ws1.column_dimensions['D'].width = 15  # H.Término
    ws1.column_dimensions['E'].width = 45  # Histórico
    
    # Título O.S.I.
    ws1.merge_cells('A1:E1')
    ws1['A1'] = f"O.S.I. {numero_os_formatado}"
    ws1['A1'].font = titulo_font
    ws1['A1'].alignment = left_align
    
    # Número da OS
    # ws1.merge_cells('A2:A2')
    # ws1['A2'] = numero_os
    # ws1['A2'].font = titulo_font
    # ws1['A2'].alignment = left_align
    
    # Cabeçalho ANDAMENTO
    ws1.merge_cells('A3:E3')
    ws1['A3'] = "ANDAMENTO"
    ws1['A3'].font = cabecalho_font
    ws1['A3'].alignment = center_align
    ws1['A3'].fill = gray_fill
    ws1['A3'].border = thin_border
    
    # Cabeçalho das colunas
    headers = ["Funcionário", "Data", "H. Início", "H.Término", "Histórico"]
    for col, header in enumerate(headers, start=1):
        cell = ws1.cell(row=4, column=col)
        cell.value = header
        cell.font = cabecalho_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Linhas para preenchimento (pontilhadas)
    for row in range(5, 50):
        for col in range(1, 6):
            cell = ws1.cell(row=row, column=col)
            if col == 2:  # Coluna Data
                cell.value = "__/__/____"
            elif col in (3, 4):  # Colunas de hora
                cell.value = "hs"
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = dotted_border
    
    # Espaço para assinaturas
    ws1.merge_cells('A51:B51')
    ws1.merge_cells('D51:E51')
    
    ws1['A51'] = "Assinatura Responsável"
    ws1['A51'].font = cabecalho_font
    ws1['A51'].alignment = center_align
    
    ws1['D51'] = "Assinatura Funcionário"
    ws1['D51'].font = cabecalho_font
    ws1['D51'].alignment = center_align
    
    # Data no rodapé
    data_atual = datetime.now().strftime("%d/%m")
    ws1['F51'] = data_atual
    ws1['F51'].font = normal_font
    ws1['F51'].alignment = right_align = Alignment(horizontal='right')
    
    # Configurações de impressão para a planilha de Andamento
    ws1.print_area = 'A1:E51'
    ws1.page_setup.orientation = 'portrait'
    ws1.page_setup.paperSize = 9  # A4
    ws1.page_setup.fitToPage = True
    ws1.page_setup.fitToHeight = 1
    ws1.page_setup.fitToWidth = 1
    ws1.print_title_rows = '1:4'  # Repetir linhas do cabeçalho em todas as páginas
    
    # Definir margens de impressão em polegadas
    ws1.page_margins.left = 0.5
    ws1.page_margins.right = 0.5
    ws1.page_margins.top = 0.5
    ws1.page_margins.bottom = 0.5
    ws1.page_margins.header = 0.3
    ws1.page_margins.footer = 0.3
    
    # Configurar cabeçalho e rodapé
    ws1.oddHeader.left.text = "OS Interna - Andamento"
    ws1.oddHeader.right.text = f"OS: {numero_os_formatado}"
    ws1.oddFooter.right.text = "Página &P de &N"
    ws1.oddFooter.left.text = datetime.now().strftime("%d/%m/%Y")
    
    # ============== PÁGINA 2: DETALHES ==============
    ws2 = wb.create_sheet("Detalhes")
    
    # Configurar larguras das colunas
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 10
    
    # Título O.S.I.
    ws2.merge_cells('A1:E1')
    ws2['A1'] = f"O.S.I. {numero_os_formatado}"
    ws2['A1'].font = titulo_font
    ws2['A1'].alignment = left_align
    
    # Número e Prazo
    # ws2['A2'] = numero_os
    # ws2['A2'].font = titulo_font
    # ws2['A2'].alignment = left_align
    
    ws2.merge_cells('B2:E2')
    ws2['B2'] = "PRAZO DE ENTREGA FINAL: __/__/____"
    ws2['B2'].font = cabecalho_font
    ws2['B2'].alignment = center_align
    
    # Informações básicas
    info_rows = [
        ("Cliente:", ""),
        ("Equipamento :", ""),
        ("Número de Série:", ""),
        ("Data de abertura O.S.I.:", ""),
        ("Horário de abertura O.S.I.:", "")
    ]
    
    for idx, (label, valor) in enumerate(info_rows, start=3):
        ws2.merge_cells(f'A{idx}:B{idx}')
        ws2.merge_cells(f'C{idx}:E{idx}')
        ws2[f'A{idx}'] = label
        ws2[f'A{idx}'].font = normal_font
        ws2[f'A{idx}'].alignment = right_align
        ws2[f'C{idx}'] = valor
        ws2[f'C{idx}'].font = normal_font
        ws2[f'C{idx}'].alignment = left_align
    
    # Tabela de tarefas
    ws2.merge_cells('A10:C10')
    ws2['A10'] = "TAREFA A EXECUTAR"
    ws2['A10'].font = cabecalho_font
    ws2['A10'].alignment = center_align
    ws2['A10'].fill = gray_fill
    ws2['A10'].border = thin_border
    
    ws2['D10'] = "PRAZO PREVISTO"
    ws2['D10'].font = cabecalho_font
    ws2['D10'].alignment = center_align
    ws2['D10'].fill = gray_fill
    ws2['D10'].border = thin_border
    
    ws2['E10'] = "OK"
    ws2['E10'].font = cabecalho_font
    ws2['E10'].alignment = center_align
    ws2['E10'].fill = gray_fill
    ws2['E10'].border = thin_border
    
    # Linhas para tarefas (vazias)
    for row in range(11, 24):
        ws2.merge_cells(f'A{row}:C{row}')
        ws2[f'A{row}'].border = dotted_border
        ws2[f'D{row}'].border = dotted_border
        ws2[f'E{row}'].border = dotted_border
    
    # Seção de observações
    ws2.merge_cells('A24:E24')
    ws2['A24'] = "OBSERVAÇÕES:"
    ws2['A24'].font = cabecalho_font
    ws2['A24'].alignment = center_align
    ws2['A24'].fill = gray_fill
    ws2['A24'].border = thin_border
    
    # Texto padrão nas observações
    ws2.merge_cells('A25:E25')
    ws2['A25'] = "Quando for trocar tubo favor:"
    ws2['A25'].font = normal_font
    ws2['A25'].alignment = left_align
    ws2['A25'].border = thin_border
    
    # Itens de observação
    observacoes = [
        "* Descrição do tubo:",
        "* N/Sº do tubo:",
        "* Tamanho do foco:"
    ]
    
    for idx, obs in enumerate(observacoes, start=26):
        ws2.merge_cells(f'A{idx}:E{idx}')
        ws2[f'A{idx}'] = obs
        ws2[f'A{idx}'].font = normal_font
        ws2[f'A{idx}'].alignment = left_align
        ws2[f'A{idx}'].border = thin_border
    
    # Ajustar a altura da célula de observações para ter espaço suficiente
    for row in range(29, 36):
        ws2.merge_cells(f'A{row}:E{row}')
        ws2[f'A{row}'].border = thin_border
    
    # Data no rodapé
    data_atual = datetime.now().strftime("%d/%m")
    ws2['E37'] = data_atual
    ws2['E37'].font = normal_font
    ws2['E37'].alignment = right_align
    
    # Configurações de impressão para a planilha de Detalhes
    ws2.print_area = 'A1:E37'
    ws2.page_setup.orientation = 'portrait'
    ws2.page_setup.paperSize = 9  # A4
    ws2.page_setup.fitToPage = True
    ws2.page_setup.fitToHeight = 1
    ws2.page_setup.fitToWidth = 1
    ws2.print_title_rows = '1:2'  # Repetir título em todas as páginas
    
    # Definir margens de impressão em polegadas
    ws2.page_margins.left = 0.5
    ws2.page_margins.right = 0.5
    ws2.page_margins.top = 0.5
    ws2.page_margins.bottom = 0.5
    ws2.page_margins.header = 0.3
    ws2.page_margins.footer = 0.3
    
    # Configurar cabeçalho e rodapé
    ws2.oddHeader.left.text = "OS Interna - Detalhes"
    ws2.oddHeader.right.text = f"OS: {numero_os_formatado}"
    ws2.oddFooter.right.text = "Página &P de &N"
    ws2.oddFooter.left.text = datetime.now().strftime("%d/%m/%Y")
    
    # ============== PÁGINA 3: MATERIAIS E PEÇAS ==============
    ws3 = wb.create_sheet("Materiais")
    
    # Configurar larguras das colunas
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 15
    
    # Título O.S.I.
    ws3.merge_cells('A1:E1')
    ws3['A1'] = f"O.S.I. {numero_os_formatado} - MATERIAIS E PEÇAS"
    ws3['A1'].font = titulo_font
    ws3['A1'].alignment = left_align
    
    # Número da OS não precisa mais aparecer na linha 2
    # ws3.merge_cells('A2:B2')
    # ws3['A2'] = numero_os
    # ws3['A2'].font = titulo_font
    # ws3['A2'].alignment = left_align
    
    # Data
    ws3.merge_cells('D2:E2')
    ws3['D2'] = "Data: __/__/____"
    ws3['D2'].font = cabecalho_font
    ws3['D2'].alignment = center_align
    
    # Cabeçalho da tabela de materiais
    headers = ["Código", "Descrição", "Quantidade", "Valor Unitário", "Valor Total"]
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=4, column=col)
        cell.value = header
        cell.font = cabecalho_font
        cell.alignment = center_align
        cell.fill = gray_fill
        cell.border = thin_border
    
    # Linhas para preenchimento de materiais
    for row in range(5, 25):
        for col in range(1, 6):
            cell = ws3.cell(row=row, column=col)
            cell.border = thin_border
            if col == 3:  # Quantidade
                cell.alignment = center_align
            elif col in (4, 5):  # Valores
                cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Totais
    ws3.merge_cells('A26:C26')
    ws3['A26'] = "TOTAL DE MATERIAIS:"
    ws3['A26'].font = cabecalho_font
    ws3['A26'].alignment = right_align
    ws3['A26'].border = thin_border
    
    ws3.merge_cells('D26:E26')
    ws3['D26'].border = thin_border
    ws3['D26'].alignment = Alignment(horizontal='right', vertical='center')
    
    # Aprovações e assinaturas
    ws3.merge_cells('A28:E28')
    ws3['A28'] = "APROVAÇÕES:"
    ws3['A28'].font = cabecalho_font
    ws3['A28'].alignment = center_align
    ws3['A28'].fill = gray_fill
    ws3['A28'].border = thin_border
    
    # Linhas para assinaturas
    assinaturas = [
        "Solicitante: _________________________________ Data: __/__/____",
        "Aprovação Técnica: __________________________ Data: __/__/____",
        "Aprovação Financeira: _______________________ Data: __/__/____"
    ]
    
    for idx, assinatura in enumerate(assinaturas, start=30):
        ws3.merge_cells(f'A{idx}:E{idx}')
        ws3[f'A{idx}'] = assinatura
        ws3[f'A{idx}'].font = normal_font
        ws3[f'A{idx}'].alignment = left_align
        ws3[f'A{idx}'].border = thin_border
    
    # Data no rodapé
    data_atual = datetime.now().strftime("%d/%m")
    ws3['E37'] = data_atual
    ws3['E37'].font = normal_font
    ws3['E37'].alignment = right_align
    
    # Configurações de impressão para a planilha de Materiais
    ws3.print_area = 'A1:E37'
    ws3.page_setup.orientation = 'portrait'
    ws3.page_setup.paperSize = 9  # A4
    ws3.page_setup.fitToPage = True
    ws3.page_setup.fitToHeight = 1
    ws3.page_setup.fitToWidth = 1
    ws3.print_title_rows = '1:4'  # Repetir cabeçalho da tabela em todas as páginas
    
    # Definir margens de impressão em polegadas
    ws3.page_margins.left = 0.5
    ws3.page_margins.right = 0.5
    ws3.page_margins.top = 0.5
    ws3.page_margins.bottom = 0.5
    ws3.page_margins.header = 0.3
    ws3.page_margins.footer = 0.3
    
    # Configurar cabeçalho e rodapé
    ws3.oddHeader.left.text = "OS Interna - Materiais e Peças"
    ws3.oddHeader.right.text = f"OS: {numero_os_formatado}"
    ws3.oddFooter.right.text = "Página &P de &N"
    ws3.oddFooter.left.text = datetime.now().strftime("%d/%m/%Y")
    
    # Salvar arquivo
    wb.save(caminho_arquivo)
    
    print(f"Arquivo salvo com sucesso em: {caminho_arquivo}")
    return caminho_arquivo

def criar_os_ziehm(numero_os, dados_cliente=None, pasta_destino=None):
    """
    Cria um arquivo Excel no formato da ZIEHM.
    
    Args:
        numero_os (str): Número da Ordem de Serviço
        dados_cliente (dict, optional): Dicionário com dados do cliente
        pasta_destino (str, optional): Caminho onde o arquivo será salvo. Se None, usa o caminho padrão.
    
    Returns:
        str: Caminho completo do arquivo salvo
    """
    # Formatação do número da OS (se ainda não estiver no formato padrão)
    if not numero_os.startswith("OS "):
        # Extrair os números do string, caso seja algo como "OSI252" ou outro formato
        numeros = re.findall(r'\d+', numero_os)
        
        if numeros:
            numero_completo = ''.join(numeros)
            # Se tiver 5 dígitos ou mais, assumimos que os 2 primeiros são do ano
            if len(numero_completo) >= 5:
                ano = numero_completo[:2]
                sequencia = numero_completo[-3:].zfill(3)  # Garantir 3 dígitos, preenchendo com zeros
            else:
                # Caso não tenha dígitos suficientes, usar o ano atual
                ano = datetime.now().strftime("%y")
                sequencia = numero_completo.zfill(3)  # Garantir 3 dígitos
                
            # Formatar no padrão correto
            numero_os_formatado = f"{ano}-{sequencia}-1"  # Formato Ziehm: 25-082-1
        else:
            # Se não tiver números, criar um padrão com o ano atual e sequência 001
            ano = datetime.now().strftime("%y")
            sequencia = "001"
            numero_os_formatado = f"{ano}-{sequencia}-1"
    else:
        # Extrair os números e reformatar para padrão Ziehm
        numeros = re.findall(r'\d+', numero_os)
        if len(numeros) >= 2:
            ano = numeros[0]
            sequencia = numeros[1].zfill(3)
            numero_os_formatado = f"{ano}-{sequencia}-1"
        else:
            # Caso não consiga extrair corretamente, usar o formato padrão
            ano = datetime.now().strftime("%y")
            sequencia = "001"
            numero_os_formatado = f"{ano}-{sequencia}-1"
            
    # Definir pasta de destino padrão se não for fornecida
    if pasta_destino is None:
        pasta_destino = os.path.join("MORACA", "MORACA1", "DOCUMENTOS", "tecnico", "andamento")
    
    # Criar diretórios se não existirem
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)
        
    # Converter o número da OS para formato usado no arquivo
    # Usar formato sem hífens para nome de arquivo
    numero_os_limpo = numero_os_formatado.replace(" ", "_").replace("/", "-").replace("-", "_")
    nome_arquivo = f"Ziehm_OS_{numero_os_limpo}.xlsx"
    caminho_arquivo = os.path.join(pasta_destino, nome_arquivo)
    
    # Verificar se o arquivo já existe
    if os.path.exists(caminho_arquivo):
        print(f"Arquivo Ziehm já existe em: {caminho_arquivo}")
        return caminho_arquivo
        
    # Se não existe, criar um novo arquivo
    print(f"Criando novo arquivo Ziehm em: {caminho_arquivo}")
    
    # Criar uma nova planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OS Ziehm"
    
    # Configurar larguras das colunas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 0.1  # Alterado de 1 para 0 para ocultar completamente a coluna
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 0.1  # Alterado de 1 para 0 para ocultar completamente a coluna
    ws.column_dimensions['I'].width = 18
    
    # Ajustar altura das linhas
    ws.row_dimensions[1].height = 25  # Linha de cabeçalho
    ws.row_dimensions[2].height = 30  # Primeira linha do logo
    ws.row_dimensions[3].height = 30  # Segunda linha do logo
    ws.row_dimensions[4].height = 30  # Terceira linha do logo
    ws.row_dimensions[5].height = 20  # Quarta linha do logo
    
    # Definir estilos
    titulo_font = Font(name='Arial', size=11, bold=True)
    cabecalho_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=9)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    
    # Bordas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Borda com linha inferior grossa para campos de assinatura
    signature_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thick')  # Borda inferior mais grossa
    )
    
    # Borda grossa para o contorno do documento
    thick_outer_border_top = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    
    thick_outer_border_bottom = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thin'),
        bottom=Side(style='thick')
    )
    
    thick_outer_border_left = Border(
        left=Side(style='thick'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_outer_border_right = Border(
        left=Side(style='thin'),
        right=Side(style='thick'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_outer_border_topleft = Border(
        left=Side(style='thick'),
        right=Side(style='thin'),
        top=Side(style='thick'),
        bottom=Side(style='thin')
    )
    
    thick_outer_border_topright = Border(
        left=Side(style='thin'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thin')
    )
    
    thick_outer_border_bottomleft = Border(
        left=Side(style='thick'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thick')
    )
    
    thick_outer_border_bottomright = Border(
        left=Side(style='thin'),
        right=Side(style='thick'),
        top=Side(style='thin'),
        bottom=Side(style='thick')
    )
    
    # Sem borda - usado para células sem borda
    sem_borda = Border(
        left=None,
        right=None,
        top=None,
        bottom=None
    )
    
    # Referência no canto direito (célula I1)
    ws['I1'] = f"Ref: ZM{ano}{sequencia}"
    ws['I1'].font = cabecalho_font
    ws['I1'].alignment = center_align
    
    # Inserir logo Ziehm - Agora cobrindo A2:E5
    logo_path = os.path.join("assets", "logoziehm.jpg")
    if os.path.exists(logo_path):
        # Mesclar células para o logo
        ws.merge_cells('A2:E5')
        
        # Adicionar o logo
        logo_img = XLImage(logo_path)
        
        # Calcular tamanho adequado para preencher a área A2:E5
        # Largura total das colunas A até E em unidades de Excel
        largura_total = sum(ws.column_dimensions[get_column_letter(i)].width for i in range(1, 6))
        
        # Altura total das linhas 2 até 5
        altura_total = sum(ws.row_dimensions[i].height for i in range(2, 6))
        
        # Converter para pixels aproximados (fator de escala)
        largura_pixels = largura_total * 7  # Aproximadamente 7 pixels por unidade de largura
        altura_pixels = altura_total * 0.75  # Converter altura de pontos para pixels (aproximado)
        
        # Ajustar tamanho da imagem
        logo_img.width = int(largura_pixels)
        logo_img.height = int(altura_pixels)
        
        # Adicionar a imagem na posição A2
        ws.add_image(logo_img, 'A2')
    
    # Cabeçalho - Telefones para Contato: - Mover para linha 6 (após o logo que ocupa até a linha 5)
    ws['F2'] = "Telefones para Contato:"
    ws['F2'].font = cabecalho_font
    ws['F2'].alignment = left_align
    ws.merge_cells('F2:H2')
    
    # Mover os telefones para baixo (linhas 7 e 8)
    ws['F3'] = "André Moraca:   (11) 2796-5727"
    ws['F3'].font = normal_font
    ws['F3'].alignment = left_align
    ws.merge_cells('F3:H3')
    
    ws['F4'] = "(11) 98525-7948"
    ws['F4'].font = normal_font
    ws['F4'].alignment = left_align
    ws.merge_cells('F4:H4')
    
    # Título da OS
    ws['A7'] = "Ordem de Serviço Número:"
    ws['A7'].font = cabecalho_font
    ws['A7'].alignment = right_align
    ws['A7'].border = thin_border
    ws.merge_cells('A7:C7')
    
    ws['D7'] = numero_os_formatado
    ws['D7'].font = cabecalho_font
    ws['D7'].alignment = center_align
    ws['D7'].border = thin_border
    
    ws['E7'] = "Data:"
    ws['E7'].font = cabecalho_font
    ws['E7'].alignment = right_align
    ws['E7'].border = thin_border
    ws.merge_cells('E7:F7')
    
    # Primeiro verificar se a célula G7 está em alguma área mesclada
    # Se estiver, desfazer a mesclagem
    merged_ranges = [r for r in ws.merged_cells.ranges if "G7" in r]
    for r in merged_ranges:
        ws.unmerge_cells(str(r))
    
    # Agora podemos atribuir o valor e depois mesclar
    ws['G7'] = formatar_data_pt_br(datetime.now())
    ws['G7'].font = normal_font
    ws['G7'].alignment = center_align
    ws['G7'].border = thin_border
    ws.merge_cells('G7:I7')
    
    # Dados do cliente
    dados_cliente = dados_cliente or {}
    
    # Nome do Cliente
    ws['A9'] = "Nome do Cliente:"
    ws['A9'].font = cabecalho_font
    ws['A9'].alignment = right_align
    ws['A9'].border = Border(left=Side(style='thin'), right=Side(style='none'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('A9:B9')
    
    # Primeiro atribuir o valor à célula C9, depois mesclar
    ws['C9'] = dados_cliente.get('cliente', '')
    ws['C9'].font = normal_font
    ws['C9'].alignment = left_align
    ws['C9'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('C9:I9')
    
    # EndereçoF
    ws['A10'] = "Endereço Instalação:"
    ws['A10'].font = cabecalho_font
    ws['A10'].alignment = right_align
    ws['A10'].border = Border(left=Side(style='thin'), right=Side(style='none'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('A10:B10')
    
    ws['C10'] = dados_cliente.get('endereco', '')
    ws['C10'].font = normal_font
    ws['C10'].alignment = left_align
    ws['C10'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('C10:I10')
    
    # Contato
    ws['A12'] = "Contato do Cliente:"
    ws['A12'].font = cabecalho_font
    ws['A12'].alignment = right_align
    ws['A12'].border = Border(left=Side(style='thin'), right=Side(style='none'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('A12:B12')
    
    ws['C12'] = dados_cliente.get('contato', '')
    ws['C12'].font = normal_font
    ws['C12'].alignment = left_align
    ws['C12'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('C12:I12')
    
    # Telefone
    ws['A13'] = "Tel:"
    ws['A13'].font = cabecalho_font
    ws['A13'].alignment = right_align
    ws['A13'].border = Border(left=Side(style='thin'), right=Side(style='none'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('A13:B13')
    
    telefone_completo = dados_cliente.get('telefone', '')
    if dados_cliente.get('telefone1'):
        if telefone_completo:
            telefone_completo += " / "
        telefone_completo += dados_cliente.get('telefone1')
    if dados_cliente.get('telefone2'):
        if telefone_completo:
            telefone_completo += " / "
        telefone_completo += dados_cliente.get('telefone2')
    
    ws['C13'] = telefone_completo
    ws['C13'].font = normal_font
    ws['C13'].alignment = left_align
    ws['C13'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('C13:I13')
    
    # Equipamento
    ws['A14'] = "Equipamento:"
    ws['A14'].font = cabecalho_font
    ws['A14'].alignment = right_align
    ws['A14'].border = Border(left=Side(style='thin'), right=Side(style='none'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('A14:B14')
    
    ws['C14'] = dados_cliente.get('equipamento', '')
    ws['C14'].font = normal_font
    ws['C14'].alignment = left_align
    ws['C14'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('C14:I14')
    
    # Modelo
    ws['A16'] = "Modelo:"
    ws['A16'].font = cabecalho_font
    ws['A16'].alignment = right_align
    ws['A16'].border = thin_border
    ws.merge_cells('A16:C16')
    
    ws['D16'] = dados_cliente.get('modelo', '')
    ws['D16'].font = normal_font
    ws['D16'].alignment = left_align
    ws['D16'].border = thin_border
    ws.merge_cells('D16:E16')
    
    # Número de Série
    ws['F16'] = "Número de Série:"
    ws['F16'].font = cabecalho_font
    ws['F16'].alignment = right_align
    ws['F16'].border = thin_border
    ws.merge_cells('F16:G16')
    
    ws['H16'] = dados_cliente.get('numero_serie', '')
    ws['H16'].font = normal_font
    ws['H16'].alignment = left_align
    ws['H16'].border = thin_border
    ws.merge_cells('H16:I16')
    
    # Tabela de registro de horas
    titles = ["Data", "Início:", "Término:", "Horas Trabalhadas:", "", "Horas de Viagem:"]
    for col, title in enumerate(titles, start=1):
        if col == 5:  # Pular uma coluna
            continue
        cell = ws.cell(row=18, column=col)
        cell.value = title
        cell.font = cabecalho_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Mesclar células de E até I nas linhas 18-24
    # Primeiro para a linha de cabeçalho (18)
    ws['E18'].value = "Horas de Viagem:"
    ws['E18'].font = cabecalho_font
    ws['E18'].alignment = center_align
    ws['E18'].border = thin_border
    ws.merge_cells('E18:I18')
    
    # Linhas para registro de horas (em branco)
    for row in range(19, 25):
        # Mesclar células de E até I para as linhas de dados
        ws.merge_cells(f'E{row}:I{row}')
        ws[f'E{row}'].border = thin_border
        
        # Tratar as demais colunas normalmente
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
        
        # Tratar a coluna 6 (Horas de Viagem) separadamente
        cell = ws.cell(row=row, column=6)
        cell.border = thin_border
    
    # Quilometragem
    # Adicionar bordas inferiores às células A25 até D25
    for col in range(1, 5):  # De A até D (1 a 4)
        cell = ws.cell(row=25, column=col)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')  # Alterado de 'thick' para 'thin'
        )
    
    ws['E25'] = "Quilometragem Total:"
    ws['E25'].font = cabecalho_font
    ws['E25'].alignment = center_align
    ws['E25'].border = thin_border
    ws.merge_cells('E25:F25')
    
    # Estender o campo "Km" até a coluna I com borda fina inferior
    ws['G25'] = "Km"
    ws['G25'].font = normal_font
    ws['G25'].alignment = center_align
    ws['G25'].border = thin_border
    ws.merge_cells('G25:I25')
    
    # Teste Inicial - Modificado para mesclar de A até I
    ws['A27'] = "Teste Inicial de Avaliação:"
    ws['A27'].font = cabecalho_font
    ws['A27'].alignment = center_align
    ws['A27'].border = thin_border
    ws.merge_cells('A27:I27')
    
    # Mesclar células A28:I30 para área de preenchimento do teste inicial
    ws['A28'] = ""  # Célula vazia para preenchimento
    ws['A28'].font = normal_font
    ws['A28'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws['A28'].border = thin_border
    ws.merge_cells('A28:I30')
    
    # Descrição dos Serviços - Modificado para mesclar de A até I
    ws['A31'] = "Descrição dos Serviços Realizados:"
    ws['A31'].font = cabecalho_font
    ws['A31'].alignment = center_align
    ws['A31'].border = thin_border
    ws.merge_cells('A31:I31')
    
    # Manutenção Preventiva
    ws['A32'] = "* Manutenção Preventiva"
    ws['A32'].font = cabecalho_font
    ws['A32'].alignment = left_align
    ws.merge_cells('A32:B32')
    
    # Área para descrição dos serviços
    # Adicionar a descrição do problema e serviço do formulário
    descricao_completa = ""
    if dados_cliente.get('descricao'):
        descricao_completa += "Descrição do problema: " + dados_cliente.get('descricao') + "\n\n"
    if dados_cliente.get('descricao_servico'):
        descricao_completa += "Serviço a realizar: " + dados_cliente.get('descricao_servico')
    
    ws['A33'] = descricao_completa
    ws['A33'].font = normal_font
    ws['A33'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws['A33'].border = thin_border
    ws.merge_cells('A33:I46')
    
    # Assinaturas
    # Configurar a altura das linhas de assinatura para 30 pontos (conforme solicitado)
    ws.row_dimensions[47].height = 30  # Ajustado para melhor espaço de assinatura
    ws.row_dimensions[49].height = 30  # Ajustado para melhor espaço de assinatura
    ws.row_dimensions[51].height = 30  # Ajustado para melhor espaço de assinatura
    ws.row_dimensions[53].height = 30  # Ajustado para melhor espaço de assinatura
    
    ws['A47'] = "Nome Legível do Cliente:"
    ws['A47'].font = cabecalho_font
    ws['A47'].alignment = left_align
    ws['A47'].border = signature_border
    ws.merge_cells('A47:C47')
    
    ws['D47'] = ""
    ws['D47'].border = signature_border
    ws.merge_cells('D47:I47')
    
    ws['A49'] = "Assinatura do Cliente:"
    ws['A49'].font = cabecalho_font
    ws['A49'].alignment = left_align
    ws['A49'].border = signature_border
    ws.merge_cells('A49:C49')
    
    ws['D49'] = ""
    ws['D49'].border = signature_border
    ws.merge_cells('D49:I49')
    
    ws['A51'] = "Nome Legível do Engenheiro/Técnico:"
    ws['A51'].font = cabecalho_font
    ws['A51'].alignment = left_align
    ws['A51'].border = signature_border
    ws.merge_cells('A51:C51')
    
    ws['D51'] = ""
    ws['D51'].border = signature_border
    ws.merge_cells('D51:I51')
    
    ws['A53'] = "Assinatura do Engenheiro/Técnico:"
    ws['A53'].font = cabecalho_font
    ws['A53'].alignment = left_align
    ws['A53'].border = signature_border
    ws.merge_cells('A53:C53')
    
    ws['D53'] = ""
    ws['D53'].border = signature_border
    ws.merge_cells('D53:I53')
    
    # Rodapé
    ws['A55'] = "Ziehm Imaging | Av. Roque Petroni Junior, 1089, Sala 904 | 04707-900 São Paulo - SP | Tel: (11) 3033-5999 |"
    ws['A55'].font = Font(name='Arial', size=8)
    ws['A55'].alignment = center_align
    ws.merge_cells('A55:I55')
    
    ws['A56'] = "Fax: (11) 3033-5997 | www.ziehm.com | vendas@ziehm.com"
    ws['A56'].font = Font(name='Arial', size=8)
    ws['A56'].alignment = center_align
    ws.merge_cells('A56:I56')
    
    # Configurações de impressão
    ws.print_area = 'A1:J56'  # Atualizado para incluir a coluna J
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    
    # Margens
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    
    # Adicionar borda grossa ao redor de todo o documento
    # Determinar a última linha e coluna com conteúdo
    max_row = 56  # A última linha com conteúdo (rodapé)
    max_col = 9   # Coluna I é a 9ª coluna
    
    # Remover todas as bordas da área de A1:I6
    for row in range(1, 7):  # Linhas 1 a 6
        for col in range(1, max_col + 1):  # Colunas A a I
            ws.cell(row=row, column=col).border = sem_borda
    
    # Aplicar bordas grossas nas extremidades do documento, começando da linha 7
    # Borda superior, agora começando na linha 7
    for col in range(1, max_col + 1):
        cell = ws.cell(row=7, column=col)
        if col == 1:  # Canto superior esquerdo (coluna A)
            cell.border = Border(
                left=Side(style=None),  # Remover borda esquerda vertical
                right=Side(style='thin'),
                top=Side(style='thick'),
                bottom=Side(style='thin')
            )
        elif col == max_col:  # Canto superior direito (coluna I)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style=None),  # Remover borda direita vertical
                top=Side(style='thick'),
                bottom=Side(style='thin')
            )
        elif col == 8:  # Coluna H
            cell.border = Border(
                left=Side(style=None),  # Remover borda esquerda vertical
                right=Side(style=None),  # Remover borda direita vertical
                top=Side(style='thick'),
                bottom=Side(style='thin')
            )
        else:  # Células intermediárias
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thick'),
                bottom=Side(style='thin')
            )
    
    # Cantos inferiores
    ws.cell(row=max_row, column=1).border = thick_outer_border_bottomleft
    ws.cell(row=max_row, column=max_col).border = thick_outer_border_bottomright
    
    # Borda inferior
    for col in range(2, max_col):
        cell = ws.cell(row=max_row, column=col)
        cell.border = thick_outer_border_bottom
    
    # Borda esquerda - começa da linha 8 para não afetar a linha 7 (nova linha superior)
    # Remover a borda esquerda completamente
    for row in range(8, max_row):
        cell = ws.cell(row=row, column=1)
        cell.border = Border(
            left=Side(style='thick'),  # Alterado para thick para adicionar borda grossa na lateral esquerda
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Borda direita - começa da linha 8 para não afetar a linha 7 (nova linha superior)
    for row in range(8, max_row):
        cell = ws.cell(row=row, column=max_col)
        cell.border = thick_outer_border_right
        
    # Remover as bordas verticais da coluna H em todas as linhas
    for row in range(7, max_row + 1):  # +1 para incluir também a última linha
        cell = ws.cell(row=row, column=8)  # Coluna H
        # Definir explicitamente a borda sem as laterais, ignorando configurações anteriores
        cell.border = Border(
            left=Side(style=None),
            right=Side(style=None),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Verificar se a célula está na linha 7 (cabeçalho) para ter borda superior espessa
        if row == 7:
            cell.border = Border(
                left=Side(style=None),
                right=Side(style=None),
                top=Side(style='thick'),
                bottom=Side(style='thin')
            )
        # Verificar se a célula está na última linha para ter borda inferior espessa
        elif row == max_row:
            cell.border = Border(
                left=Side(style=None),
                right=Side(style=None),
                top=Side(style='thin'),
                bottom=Side(style='thick')
            )
            
    # Adicionando célula na coluna J para completar a volta da borda direita
    ws.column_dimensions['J'].width = 0.1  # Coluna invisível, apenas para a borda
    for row in range(7, max_row + 1):
        cell = ws.cell(row=row, column=10)  # Coluna J (após a coluna I)
        if row == 7:  # Primeira linha com borda superior espessa
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thin')
            )
        elif row == max_row:  # Última linha com borda inferior espessa
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thick'),
                top=Side(style='thin'),
                bottom=Side(style='thick')
            )
        else:  # Linhas intermediárias
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thick'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Salvar arquivo
    wb.save(caminho_arquivo)
    
    print(f"Arquivo OS Ziehm salvo com sucesso em: {caminho_arquivo}")
    return caminho_arquivo

if __name__ == "__main__":
    # Exemplo de uso
    numero_os = "25019"  # Substituído pelo formato correto durante a execução
    pasta_destino = os.path.join("/home/gabriel-balsarin/Projetos/Moraca/MORACA", "MORACA1", "DOCUMENTOS", "tecnico", "andamento")
    criar_os_interna_andamento(numero_os, pasta_destino) 